"""
Motor de detecção de fotos duplicadas.

Fluxo:
1. Lê a planilha (.xlsx) e extrai os links de imagem das colunas de fotos
   (por padrão D, E e F). Cada célula pode conter vários links (um por
   linha dentro da célula).
2. Baixa cada imagem e calcula o hash SHA-256 dos bytes (duplicata = mesmo arquivo).
   Atalho: se dois links têm o mesmo ID de arquivo na URL, já são a mesma imagem.
3. Agrupa as imagens idênticas. A 1ª ocorrência (linha/coluna mais acima) é a
   "original"; as seguintes são "duplicadas".
4. Gera um novo Excel colorindo exatamente as células duplicadas e adiciona
   colunas indicando de qual linha/coluna (original) cada duplicata foi copiada.
"""

import re
import hashlib
from dataclasses import dataclass, field
from concurrent.futures import ThreadPoolExecutor, as_completed

import requests
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ------------------------------------------------------------------ #
# Configuração
# ------------------------------------------------------------------ #

# Colunas onde ficam os links das fotos (D=4, E=5, F=6). Ajustável.
FOTOS_COLS = [4, 5, 6]
HEADER_ROW = 1
DOWNLOAD_TIMEOUT = 20      # segundos por imagem
MAX_WORKERS = 8            # downloads paralelos

# Cabeçalhos enviados no download. Se os links do Sebrae passarem a exigir
# login, basta preencher AUTH_HEADERS com o token/cookie de sessão.
DEFAULT_HEADERS = {
    "User-Agent": "Mozilla/5.0 (compatible; VerificadorFotosSebrae/1.0)"
}
AUTH_HEADERS = {}          # ex.: {"Authorization": "Bearer <token>"}

# Extrai o ID do arquivo de URLs no formato .../arquivo/<ID>/binario
_ID_RE = re.compile(r"/arquivo/([0-9a-fA-F]+)/", re.IGNORECASE)


# ------------------------------------------------------------------ #
# Estruturas de dados
# ------------------------------------------------------------------ #

@dataclass
class Foto:
    row: int                    # linha na planilha (1-indexed)
    col: int                     # coluna na planilha (1-indexed; D=4, E=5, F=6)
    idx_na_celula: int          # posição do link dentro da célula (0,1,2...)
    url: str
    file_id: str = ""           # ID extraído da URL, se houver
    sha256: str = ""            # hash do conteúdo baixado
    erro: str = ""              # mensagem se o download falhou
    tamanho: int = 0            # bytes

    @property
    def col_letra(self):
        return get_column_letter(self.col)


@dataclass
class Duplicata:
    foto: Foto                  # a foto duplicada
    original: Foto              # a primeira ocorrência (original)


@dataclass
class Resultado:
    fotos: list = field(default_factory=list)
    duplicatas: list = field(default_factory=list)   # list[Duplicata]
    falhas: list = field(default_factory=list)       # list[Foto] com erro
    total_links: int = 0
    total_baixadas: int = 0


# ------------------------------------------------------------------ #
# 1. Leitura da planilha
# ------------------------------------------------------------------ #

def _split_links(valor):
    """Uma célula pode ter vários links (separados por quebra de linha,
    espaço, vírgula ou ponto-e-vírgula). Retorna lista limpa."""
    if valor is None:
        return []
    texto = str(valor)
    partes = re.split(r"[\n\r,; ]+", texto)
    return [p.strip() for p in partes if p.strip().lower().startswith("http")]


def extrair_fotos(caminho_xlsx, fotos_cols=FOTOS_COLS):
    """Lê a planilha e retorna a lista de Foto (uma por link encontrado),
    varrendo TODAS as colunas informadas (padrão: D, E e F) em cada linha."""
    wb = load_workbook(caminho_xlsx)
    ws = wb.active
    fotos = []
    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        for col in fotos_cols:
            valor = ws.cell(row=row, column=col).value
            links = _split_links(valor)
            for idx, url in enumerate(links):
                m = _ID_RE.search(url)
                file_id = m.group(1) if m else ""
                fotos.append(Foto(row=row, col=col, idx_na_celula=idx,
                                   url=url, file_id=file_id))
    return fotos


# ------------------------------------------------------------------ #
# 2. Download + hash
# ------------------------------------------------------------------ #

def _baixar_e_hashear(foto: Foto):
    headers = {**DEFAULT_HEADERS, **AUTH_HEADERS}
    try:
        resp = requests.get(foto.url, headers=headers,
                            timeout=DOWNLOAD_TIMEOUT, stream=True)
        if resp.status_code != 200:
            foto.erro = f"HTTP {resp.status_code}"
            return foto
        conteudo = resp.content
        if not conteudo:
            foto.erro = "arquivo vazio"
            return foto
        foto.sha256 = hashlib.sha256(conteudo).hexdigest()
        foto.tamanho = len(conteudo)
    except requests.RequestException as e:
        foto.erro = f"falha de rede: {type(e).__name__}"
    return foto


def baixar_todas(fotos, progress_cb=None):
    """Baixa todas as fotos em paralelo. progress_cb(feito, total) opcional."""
    total = len(fotos)
    feito = 0
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
        futuros = {ex.submit(_baixar_e_hashear, f): f for f in fotos}
        for fut in as_completed(futuros):
            fut.result()
            feito += 1
            if progress_cb:
                progress_cb(feito, total)
    return fotos


# ------------------------------------------------------------------ #
# 3. Detecção de duplicatas
# ------------------------------------------------------------------ #

def detectar(fotos, ignorar_mesma_linha=True):
    """Agrupa fotos idênticas. A 1ª ocorrência é a original.
    Considera idênticas se: mesmo sha256 OU (sem hash) mesmo file_id.

    ignorar_mesma_linha: se True, duas fotos iguais no MESMO atendimento
    (mesma linha, independente da coluna D/E/F) não são tratadas como
    duplicata — provável reenvio acidental, não uso indevido entre
    atendimentos distintos."""
    resultado = Resultado(fotos=fotos, total_links=len(fotos))

    # ordena por linha, coluna e posição para que "original" seja sempre a de cima
    fotos_ok = sorted(
        [f for f in fotos if f.sha256],
        key=lambda f: (f.row, f.col, f.idx_na_celula)
    )
    resultado.total_baixadas = len(fotos_ok)
    resultado.falhas = [f for f in fotos if not f.sha256]

    visto = {}   # chave -> Foto original
    for f in fotos_ok:
        chave = f.sha256
        if chave in visto:
            orig = visto[chave]
            if ignorar_mesma_linha and orig.row == f.row:
                continue
            resultado.duplicatas.append(Duplicata(foto=f, original=orig))
        else:
            visto[chave] = f

    # Atalho por file_id, para pares que não puderam ser baixados mas têm ID igual.
    # (só marca se ainda não foi pego pelo hash)
    ja_dup = {(d.foto.row, d.foto.col, d.foto.idx_na_celula) for d in resultado.duplicatas}
    visto_id = {}
    for f in sorted(fotos, key=lambda f: (f.row, f.col, f.idx_na_celula)):
        if not f.file_id:
            continue
        if f.file_id in visto_id:
            orig = visto_id[f.file_id]
            if ignorar_mesma_linha and orig.row == f.row:
                continue
            if (f.row, f.col, f.idx_na_celula) not in ja_dup and f.sha256 == "":
                resultado.duplicatas.append(
                    Duplicata(foto=f, original=orig))
        else:
            visto_id[f.file_id] = f

    return resultado


# ------------------------------------------------------------------ #
# 4. Geração do Excel de saída
# ------------------------------------------------------------------ #

FILL_DUP = PatternFill("solid", fgColor="FFC7CE")      # vermelho claro (duplicata)
FILL_ORIG = PatternFill("solid", fgColor="FFEB9C")     # amarelo (original c/ cópia)
FILL_ERRO = PatternFill("solid", fgColor="D9D9D9")     # cinza (falha download)
FONT_DUP = Font(color="9C0006")
FONT_HDR = Font(bold=True, color="FFFFFF")
FILL_HDR = PatternFill("solid", fgColor="1F4E78")


def gerar_excel(caminho_entrada, resultado: Resultado, caminho_saida,
                fotos_cols=FOTOS_COLS):
    """Cria um novo xlsx com duplicatas coloridas célula a célula (D/E/F)
    e adiciona colunas de diagnóstico."""
    wb = load_workbook(caminho_entrada)
    ws = wb.active

    # índice por CÉLULA exata (linha, coluna), não só pela linha
    dups_por_celula = {}
    for d in resultado.duplicatas:
        dups_por_celula.setdefault((d.foto.row, d.foto.col), []).append(d)
    originais_por_celula = {(d.original.row, d.original.col) for d in resultado.duplicatas}
    falhas_por_celula = {}
    for f in resultado.falhas:
        falhas_por_celula.setdefault((f.row, f.col), []).append(f)

    # colunas novas de diagnóstico
    col_status = ws.max_column + 1
    col_detalhe = ws.max_column + 2
    ws.cell(row=HEADER_ROW, column=col_status, value="Status Verificação")
    ws.cell(row=HEADER_ROW, column=col_detalhe, value="Detalhe / Original")
    for c in (col_status, col_detalhe):
        cel = ws.cell(row=HEADER_ROW, column=c)
        cel.font = FONT_HDR
        cel.fill = FILL_HDR
        cel.alignment = Alignment(horizontal="center", vertical="center")

    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        partes = []
        status = "OK"

        for col in fotos_cols:
            chave = (row, col)
            cel_foto = ws.cell(row=row, column=col)

            if chave in dups_por_celula:
                status = "DUPLICADA"
                cel_foto.fill = FILL_DUP
                cel_foto.font = FONT_DUP
                for d in dups_por_celula[chave]:
                    partes.append(
                        f"{d.foto.col_letra}{d.foto.row} é cópia de "
                        f"{d.original.col_letra}{d.original.row} "
                        f"({_rotulo(ws, d.original.row)})"
                    )
            elif chave in originais_por_celula:
                if status != "DUPLICADA":
                    status = "ORIGINAL (tem cópias)"
                cel_foto.fill = FILL_ORIG

            if chave in falhas_por_celula:
                if status == "OK":
                    status = "NÃO VERIFICADA"
                if chave not in dups_por_celula and chave not in originais_por_celula:
                    cel_foto.fill = FILL_ERRO
                for f in falhas_por_celula[chave]:
                    partes.append(
                        f"{f.col_letra}{f.row}: falha no download ({f.erro})")

        ws.cell(row=row, column=col_status, value=status)
        ws.cell(row=row, column=col_detalhe, value="  |  ".join(partes))

    ws.column_dimensions[get_column_letter(col_status)].width = 22
    ws.column_dimensions[get_column_letter(col_detalhe)].width = 70

    # aba de resumo
    _aba_resumo(wb, resultado, ws)

    wb.save(caminho_saida)
    return caminho_saida


def _rotulo(ws, row):
    """Monta um rótulo curto da linha (razão social, se existir na col C)."""
    razao = ws.cell(row=row, column=3).value
    return str(razao)[:40] if razao else f"linha {row}"


def _aba_resumo(wb, resultado: Resultado, ws_orig):
    ws = wb.create_sheet("Resumo Duplicatas")
    headers = ["Célula Duplicada", "Linha Duplicada", "Coluna Duplicada",
               "Célula Original", "Linha da Original", "Coluna Original",
               "Razão Social (dup)", "Razão Social (original)", "URL Duplicada"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cel = ws.cell(row=1, column=c)
        cel.font = FONT_HDR
        cel.fill = FILL_HDR
    for d in sorted(resultado.duplicatas, key=lambda x: (x.foto.row, x.foto.col)):
        ws.append([
            f"{d.foto.col_letra}{d.foto.row}",
            d.foto.row,
            d.foto.col_letra,
            f"{d.original.col_letra}{d.original.row}",
            d.original.row,
            d.original.col_letra,
            ws_orig.cell(row=d.foto.row, column=3).value,
            ws_orig.cell(row=d.original.row, column=3).value,
            d.foto.url,
        ])
    widths = [16, 14, 14, 16, 16, 14, 35, 35, 70]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ------------------------------------------------------------------ #
# Orquestração completa (usada pela app web)
# ------------------------------------------------------------------ #

def processar(caminho_entrada, caminho_saida, progress_cb=None, fotos_cols=FOTOS_COLS,
              ignorar_mesma_linha=True):
    fotos = extrair_fotos(caminho_entrada, fotos_cols=fotos_cols)
    baixar_todas(fotos, progress_cb=progress_cb)
    resultado = detectar(fotos, ignorar_mesma_linha=ignorar_mesma_linha)
    gerar_excel(caminho_entrada, resultado, caminho_saida, fotos_cols=fotos_cols)
    return resultado
